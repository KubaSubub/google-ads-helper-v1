"""CSV/Excel export endpoints."""

import io
from datetime import date, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import AdGroup, Campaign, Client, Keyword, MetricDaily, SearchTerm
from app.models.action_log import ActionLog
from app.utils.quality_score import build_subcomponent_issues as _build_subcomponent_issues, get_primary_issue as _get_primary_issue, build_recommendation as _build_recommendation
from app.services.recommendations import recommendations_engine
from app.utils.formatters import micros_to_currency

_SUBCOMP_DISPLAY = {1: "Poniżej średniej", 2: "Średnia", 3: "Powyżej średniej"}

router = APIRouter(prefix="/export", tags=["Export"])


def _validate_client(db: Session, client_id: int) -> Client:
    """Validate client exists, raise 404 if not."""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail=f"Client {client_id} not found")
    return client


def _validate_campaign(db: Session, campaign_id: int) -> Campaign:
    """Validate campaign exists, raise 404 if not."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail=f"Campaign {campaign_id} not found")
    return campaign


@router.get("/search-terms")
def export_search_terms(
    client_id: int = Query(...),
    format: str = Query("xlsx", description="xlsx or csv"),
    db: Session = Depends(get_db),
):
    """Export all search terms for a client as XLSX or CSV."""
    _validate_client(db, client_id)

    terms = (
        db.query(SearchTerm)
        .outerjoin(AdGroup, SearchTerm.ad_group_id == AdGroup.id)
        .outerjoin(Campaign, AdGroup.campaign_id == Campaign.id)
        .filter(
            (Campaign.client_id == client_id)
            | (SearchTerm.campaign_id.in_(
                db.query(Campaign.id).filter(Campaign.client_id == client_id)
            ))
        )
        .order_by(SearchTerm.cost_micros.desc())
        .all()
    )

    headers = ["Fraza", "Klikniecia", "Wyswietlenia", "Koszt (USD)", "Konwersje", "CTR %", "CPC (USD)", "Koszt/konw. (USD)"]

    def _row(term):
        cost_usd = micros_to_currency(term.cost_micros)
        cpc = cost_usd / term.clicks if term.clicks else 0
        cpconv = cost_usd / term.conversions if (term.conversions or 0) > 0 else 0
        ctr_pct = round(term.ctr or 0, 2)
        return [term.text, term.clicks or 0, term.impressions or 0, cost_usd, term.conversions or 0, ctr_pct, round(cpc, 2), round(cpconv, 2)]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for term in terms:
            row = _row(term)
            output.write(f'"{row[0]}",{row[1]},{row[2]},{row[3]},{row[4]},{row[5]},{row[6]},{row[7]}\n')
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=search_terms.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Terms"
    ws.append(headers)
    for term in terms:
        ws.append(_row(term))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=search_terms.xlsx"},
    )


@router.get("/metrics")
def export_metrics(
    campaign_id: int = Query(...),
    days: int = Query(30, ge=1, le=365),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
):
    """Export daily metrics for a campaign as XLSX or CSV."""
    _validate_campaign(db, campaign_id)

    cutoff = date.today() - timedelta(days=days)
    metrics = (
        db.query(MetricDaily)
        .filter(MetricDaily.campaign_id == campaign_id, MetricDaily.date >= cutoff)
        .order_by(MetricDaily.date)
        .all()
    )

    headers = ["Data", "Klikniecia", "Wyswietlenia", "CTR %", "Koszt (USD)", "Konwersje", "CPA (USD)", "ROAS", "Avg CPC (USD)"]

    def _row(metric):
        cost_usd = micros_to_currency(metric.cost_micros)
        cpa = cost_usd / metric.conversions if (metric.conversions or 0) > 0 else 0
        ctr_pct = round(metric.ctr or 0, 2)
        avg_cpc_usd = micros_to_currency(metric.avg_cpc_micros)
        return [str(metric.date), metric.clicks or 0, metric.impressions or 0, ctr_pct, cost_usd, metric.conversions or 0, round(cpa, 2), round(metric.roas or 0, 2), avg_cpc_usd]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for metric in metrics:
            row = _row(metric)
            output.write(",".join(str(value) for value in row) + "\n")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=metrics_campaign_{campaign_id}.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Metrics"
    ws.append(headers)
    for metric in metrics:
        ws.append(_row(metric))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=metrics_campaign_{campaign_id}.xlsx"},
    )


@router.get("/keywords")
def export_keywords(
    client_id: int = Query(...),
    campaign_id: int = Query(None),
    include_removed: bool = Query(False),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
):
    """Export keywords for a client as XLSX or CSV."""
    _validate_client(db, client_id)
    if campaign_id is not None:
        _validate_campaign(db, campaign_id)

    keywords = (
        db.query(Keyword, Campaign.name.label("campaign_name"), AdGroup.name.label("ad_group_name"))
        .join(AdGroup, Keyword.ad_group_id == AdGroup.id)
        .join(Campaign, AdGroup.campaign_id == Campaign.id)
        .filter(Campaign.client_id == client_id)
    )
    if campaign_id is not None:
        keywords = keywords.filter(Campaign.id == campaign_id)
    if not include_removed:
        keywords = keywords.filter(Keyword.status.in_(("ENABLED", "PAUSED")))
    keywords = keywords.order_by(Keyword.cost_micros.desc()).all()

    headers = ["Kampania", "Grupa reklam", "Slowo kluczowe", "Dopasowanie", "Status", "Klikniecia", "Wyswietlenia", "Koszt (USD)", "Konwersje", "CTR %", "Avg CPC (USD)"]

    def _row(keyword, campaign_name, ad_group_name):
        cost_usd = micros_to_currency(keyword.cost_micros)
        ctr_pct = round(keyword.ctr or 0, 2)
        avg_cpc_usd = micros_to_currency(keyword.avg_cpc_micros)
        return [campaign_name, ad_group_name, keyword.text, keyword.match_type, keyword.status, keyword.clicks or 0, keyword.impressions or 0, cost_usd, keyword.conversions or 0, ctr_pct, avg_cpc_usd]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for keyword, campaign_name, ad_group_name in keywords:
            row = _row(keyword, campaign_name, ad_group_name)
            output.write(",".join(f'"{value}"' if isinstance(value, str) else str(value) for value in row) + "\n")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=keywords.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords"
    ws.append(headers)
    for keyword, campaign_name, ad_group_name in keywords:
        ws.append(_row(keyword, campaign_name, ad_group_name))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=keywords.xlsx"},
    )


@router.get("/recommendations")
def export_recommendations(
    client_id: int = Query(...),
    days: int = Query(30, ge=1, le=365),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
):
    """Export recommendations for a client as XLSX or CSV."""
    _validate_client(db, client_id)

    results = recommendations_engine.generate_all(db, client_id, days)

    headers = ["Priorytet", "Typ", "Encja", "Kampania", "Powod", "Obecna wartosc", "Rekomendacja", "Szacowany wplyw"]

    def _row(result):
        return [
            result.get("priority", ""),
            result.get("type", ""),
            result.get("entity_name", ""),
            result.get("campaign_name", ""),
            result.get("reason", ""),
            result.get("current_value", ""),
            result.get("recommended_action", ""),
            result.get("estimated_impact", ""),
        ]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for result in results:
            row = _row(result)
            output.write(",".join(f'"{value}"' for value in row) + "\n")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=recommendations.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    ws.append(headers)
    for result in results:
        ws.append(_row(result))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=recommendations.xlsx"},
    )


@router.get("/quality-score")
def export_quality_score(
    client_id: int = Query(...),
    qs_threshold: int = Query(5, ge=1, le=10),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
):
    """Export Quality Score audit as XLSX or CSV."""
    _validate_client(db, client_id)

    keywords = (
        db.query(Keyword, Campaign.name.label("campaign_name"), AdGroup.name.label("ad_group_name"))
        .join(AdGroup, Keyword.ad_group_id == AdGroup.id)
        .join(Campaign, AdGroup.campaign_id == Campaign.id)
        .filter(
            Campaign.client_id == client_id,
            Keyword.status == "ENABLED",
            Keyword.quality_score.isnot(None),
            Keyword.quality_score > 0,
        )
        .order_by(Keyword.quality_score.asc())
        .all()
    )

    headers = [
        "Słowo kluczowe", "Kampania", "Grupa reklam", "Dopasowanie", "QS",
        "Oczekiwany CTR", "Trafność reklamy", "Strona docelowa",
        "CTR %", "Kliknięcia", "Wyświetlenia", "Koszt (zł)", "Konwersje",
        "IS utracony (ranking)", "Główny problem", "Rekomendacja",
    ]

    def _row(kw, campaign_name, ad_group_name):
        cost = micros_to_currency(kw.cost_micros)
        primary = _get_primary_issue(kw)
        issue_labels = {
            "expected_ctr": "Oczekiwany CTR",
            "ad_relevance": "Trafność reklamy",
            "landing_page": "Strona docelowa",
        }
        return [
            kw.text,
            campaign_name,
            ad_group_name,
            kw.match_type or "",
            kw.quality_score,
            _SUBCOMP_DISPLAY.get(kw.historical_search_predicted_ctr, "—"),
            _SUBCOMP_DISPLAY.get(kw.historical_creative_quality, "—"),
            _SUBCOMP_DISPLAY.get(kw.historical_landing_page_quality, "—"),
            round(kw.ctr or 0, 2),
            kw.clicks or 0,
            kw.impressions or 0,
            round(cost, 2),
            round(kw.conversions or 0, 1),
            f"{round((kw.search_rank_lost_is or 0) * 100, 1)}%",
            issue_labels.get(primary, "—"),
            _build_recommendation(kw),
        ]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for kw, campaign_name, ad_group_name in keywords:
            row = _row(kw, campaign_name, ad_group_name)
            output.write(",".join(f'"{v}"' if isinstance(v, str) else str(v) for v in row) + "\n")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=quality_score_audit.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Score Audit"
    ws.append(headers)
    for kw, campaign_name, ad_group_name in keywords:
        ws.append(_row(kw, campaign_name, ad_group_name))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quality_score_audit.xlsx"},
    )


@router.get("/actions")
def export_actions(
    client_id: int = Query(...),
    format: str = Query("csv", description="csv or xlsx"),
    db: Session = Depends(get_db),
):
    """Export action history for a client as CSV or XLSX."""
    _validate_client(db, client_id)

    actions = (
        db.query(ActionLog)
        .filter(ActionLog.client_id == client_id)
        .order_by(ActionLog.executed_at.desc())
        .all()
    )

    # Batch-load entity names
    kw_ids = set()
    camp_ids = set()
    for a in actions:
        try:
            eid = int(a.entity_id) if a.entity_id else None
        except (ValueError, TypeError):
            continue
        if eid and a.entity_type == "keyword":
            kw_ids.add(eid)
        elif eid and a.entity_type == "campaign":
            camp_ids.add(eid)

    kw_map = {}
    ag_map = {}
    if kw_ids:
        keywords = db.query(Keyword).filter(Keyword.id.in_(kw_ids)).all()
        kw_map = {k.id: k for k in keywords}
        ag_ids = {k.ad_group_id for k in keywords if k.ad_group_id}
        if ag_ids:
            ad_groups = db.query(AdGroup).filter(AdGroup.id.in_(ag_ids)).all()
            ag_map = {ag.id: ag for ag in ad_groups}
            camp_ids.update(ag.campaign_id for ag in ad_groups if ag.campaign_id)

    camp_map = {}
    if camp_ids:
        campaigns = db.query(Campaign).filter(Campaign.id.in_(camp_ids)).all()
        camp_map = {c.id: c for c in campaigns}

    OP_LABELS = {
        "PAUSE_KEYWORD": "Wstrzymano keyword",
        "ENABLE_KEYWORD": "Włączono keyword",
        "UPDATE_BID": "Zmieniono stawkę",
        "SET_KEYWORD_BID": "Przywrócono stawkę",
        "ADD_KEYWORD": "Dodano keyword",
        "ADD_NEGATIVE": "Dodano negative",
        "PAUSE_AD": "Wstrzymano reklamę",
        "INCREASE_BUDGET": "Zwiększono budżet",
        "SET_BUDGET": "Przywrócono budżet",
        "DECREASE_BUDGET": "Zmniejszono budżet",
    }

    headers = ["Data", "Akcja", "Typ akcji", "Encja", "Kampania", "Status", "Stara wartość", "Nowa wartość"]

    def _resolve_entity(action):
        entity_name = ""
        campaign_name = ""
        try:
            eid = int(action.entity_id) if action.entity_id else None
        except (ValueError, TypeError):
            return entity_name, campaign_name
        if not eid:
            return entity_name, campaign_name
        if action.entity_type == "keyword":
            kw = kw_map.get(eid)
            if kw:
                entity_name = kw.text or ""
                ag = ag_map.get(kw.ad_group_id)
                if ag:
                    camp = camp_map.get(ag.campaign_id)
                    if camp:
                        campaign_name = camp.name or ""
        elif action.entity_type == "campaign":
            camp = camp_map.get(eid)
            if camp:
                entity_name = camp.name or ""
                campaign_name = camp.name or ""
        return entity_name, campaign_name

    def _row(action):
        entity_name, campaign_name = _resolve_entity(action)
        import json
        old_val = ""
        new_val = ""
        if action.old_value_json:
            try:
                old_val = json.dumps(action.old_value_json, ensure_ascii=False) if isinstance(action.old_value_json, dict) else str(action.old_value_json)
            except Exception:
                old_val = str(action.old_value_json)
        if action.new_value_json:
            try:
                new_val = json.dumps(action.new_value_json, ensure_ascii=False) if isinstance(action.new_value_json, dict) else str(action.new_value_json)
            except Exception:
                new_val = str(action.new_value_json)
        return [
            str(action.executed_at) if action.executed_at else "",
            OP_LABELS.get(action.action_type, action.action_type or ""),
            action.action_type or "",
            entity_name,
            campaign_name,
            action.status or "",
            old_val,
            new_val,
        ]

    if format == "csv":
        output = io.StringIO()
        output.write(",".join(headers) + "\n")
        for action in actions:
            row = _row(action)
            output.write(",".join(f'"{str(v).replace(chr(34), chr(39))}"' for v in row) + "\n")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=action_history.csv"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Action History"
    ws.append(headers)
    for action in actions:
        ws.append(_row(action))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=action_history.xlsx"},
    )
