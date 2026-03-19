"""Tests for export endpoints — XLSX and CSV downloads."""

import io
from datetime import date, timedelta, timezone

import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app.database import get_db
from app.main import app
from app.models.ad_group import AdGroup
from app.models.campaign import Campaign
from app.models.client import Client
from app.models.keyword import Keyword
from app.models.metric_daily import MetricDaily
from app.models.search_term import SearchTerm


@pytest.fixture
def api_client(db):
    def _override():
        yield db

    app.dependency_overrides[get_db] = _override
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.pop(get_db, None)


def _seed_search_terms(db):
    client = Client(name="Export Client", google_customer_id="1112223330")
    db.add(client)
    db.flush()

    campaign = Campaign(client_id=client.id, google_campaign_id="ec1", name="Camp A", status="ENABLED", campaign_type="SEARCH")
    db.add(campaign)
    db.flush()

    ad_group = AdGroup(campaign_id=campaign.id, google_ad_group_id="ag1", name="Group A", status="ENABLED")
    db.add(ad_group)
    db.flush()

    today = date.today()
    week_ago = today - timedelta(days=7)

    db.add(SearchTerm(
        ad_group_id=ad_group.id,
        text="buty sportowe",
        clicks=50,
        impressions=500,
        cost_micros=10_000_000,
        conversions=3.0,
        ctr=10.0,  # 10% as percentage
        date_from=week_ago,
        date_to=today,
    ))
    db.add(SearchTerm(
        ad_group_id=ad_group.id,
        text="buty zimowe",
        clicks=0,
        impressions=100,
        cost_micros=0,
        conversions=0,
        ctr=0,
        date_from=week_ago,
        date_to=today,
    ))
    db.commit()
    return client


def _seed_keywords(db):
    client = Client(name="KW Export", google_customer_id="2223334440")
    db.add(client)
    db.flush()

    campaign = Campaign(client_id=client.id, google_campaign_id="kc1", name="KW Camp", status="ENABLED", campaign_type="SEARCH")
    db.add(campaign)
    db.flush()

    ad_group = AdGroup(campaign_id=campaign.id, google_ad_group_id="kag1", name="KW Group", status="ENABLED")
    db.add(ad_group)
    db.flush()

    db.add(Keyword(
        ad_group_id=ad_group.id,
        google_keyword_id="kw1",
        text="buty nike",
        match_type="EXACT",
        status="ENABLED",
        clicks=20,
        impressions=200,
        cost_micros=5_000_000,
        conversions=1.0,
        ctr=10.0,  # 10% as percentage
        avg_cpc_micros=250_000,
    ))
    db.add(Keyword(
        ad_group_id=ad_group.id,
        google_keyword_id="kw2",
        text="stare buty",
        match_type="BROAD",
        status="REMOVED",
        clicks=0,
        impressions=0,
        cost_micros=0,
        conversions=0,
        ctr=0,
        avg_cpc_micros=0,
    ))
    db.commit()
    return client


# ---------------------------------------------------------------------------
# Search Terms export
# ---------------------------------------------------------------------------

class TestExportSearchTerms:
    def test_xlsx_returns_valid_workbook(self, api_client, db):
        client = _seed_search_terms(db)

        resp = api_client.get(f"/api/v1/export/search-terms?client_id={client.id}&format=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "Fraza"  # header
        assert len(rows) == 3  # header + 2 terms

    def test_csv_returns_valid_data(self, api_client, db):
        client = _seed_search_terms(db)

        resp = api_client.get(f"/api/v1/export/search-terms?client_id={client.id}&format=csv")
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]

        # Remove BOM if present
        text = resp.content.decode("utf-8-sig")
        lines = text.strip().split("\n")
        assert len(lines) == 3  # header + 2 terms
        assert "buty sportowe" in lines[1]

    def test_404_for_nonexistent_client(self, api_client, db):
        resp = api_client.get("/api/v1/export/search-terms?client_id=99999&format=xlsx")
        assert resp.status_code == 404

    def test_empty_terms_returns_header_only_xlsx(self, api_client, db):
        client = Client(name="Empty", google_customer_id="3334445550")
        db.add(client)
        db.commit()

        resp = api_client.get(f"/api/v1/export/search-terms?client_id={client.id}&format=xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 1  # header only


# ---------------------------------------------------------------------------
# Keywords export
# ---------------------------------------------------------------------------

class TestExportKeywords:
    def test_xlsx_excludes_removed_by_default(self, api_client, db):
        client = _seed_keywords(db)

        resp = api_client.get(f"/api/v1/export/keywords?client_id={client.id}&format=xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2  # header + 1 enabled keyword
        assert rows[1][2] == "buty nike"

    def test_xlsx_includes_removed_when_requested(self, api_client, db):
        client = _seed_keywords(db)

        resp = api_client.get(f"/api/v1/export/keywords?client_id={client.id}&include_removed=true&format=xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 3  # header + 2 keywords

    def test_csv_format(self, api_client, db):
        client = _seed_keywords(db)

        resp = api_client.get(f"/api/v1/export/keywords?client_id={client.id}&format=csv")
        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]


# ---------------------------------------------------------------------------
# Metrics export
# ---------------------------------------------------------------------------

class TestExportMetrics:
    def test_xlsx_returns_daily_rows(self, api_client, db):
        client = Client(name="Metrics", google_customer_id="4445556660")
        db.add(client)
        db.flush()

        campaign = Campaign(client_id=client.id, google_campaign_id="mc1", name="Metrics Camp", status="ENABLED", campaign_type="SEARCH")
        db.add(campaign)
        db.flush()

        today = date.today()
        for i in range(3):
            db.add(MetricDaily(
                campaign_id=campaign.id,
                date=today - timedelta(days=i),
                clicks=10,
                impressions=100,
                cost_micros=1_000_000,
                conversions=1.0,
                conversion_value_micros=5_000_000,
            ))
        db.commit()

        resp = api_client.get(f"/api/v1/export/metrics?campaign_id={campaign.id}&days=7&format=xlsx")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[0][0] == "Data"
        assert len(rows) == 4  # header + 3 days

    def test_404_for_nonexistent_campaign(self, api_client, db):
        resp = api_client.get("/api/v1/export/metrics?campaign_id=99999&format=xlsx")
        assert resp.status_code == 404
